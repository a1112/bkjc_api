from pathlib import Path

import openpyxl
import config


from subbProject.SteelLevel.Base.module import SteelProject
from subbProject.SteelLevel.DataSet.LevelSet import getUserDatByPackageNo
from subbProject.SteelLevel.DataSet.models import UserData, SteelLevel

def level_to_str(level_int):
    if level_int==1:
        return '一等品'
    elif level_int==2:
        return '二等品'
    elif level_int==3:
        return '三等品'
    return ""

def saveExcel(data, excel_path):
    print("saveExcel")
    workbook = openpyxl.load_workbook(config.steelLevelTemplateDataOut)
    ws1 = workbook.get_sheet_by_name("Sheet1")
    workbook.active = 0  # 设置active参数，即工作表索引值
    for colIndex, dataItem in enumerate(data):
        dataItem: SteelProject
        useData = getUserDatByPackageNo(dataItem.packageName)
        useCode = ""
        useNode = ""
        if useData:
            useData = useData[0]
            useData : UserData
            useCode = useData.steelLevel
            useNode = useData.scrapReasonCode

        if isinstance(dataItem,SteelLevel):
            for rowIndex, value in enumerate([
                dataItem.seqNo,
                config.plant_classification,
                config.productionLine_classification,
                dataItem.steelName,
                dataItem.packageName,
                dataItem.steelType,
                "",
                dataItem.length,
                dataItem.width,
                dataItem.thick,
                dataItem.detectTime,
                dataItem.levelInfo_up,
                dataItem.levelInfo_under,
                level_to_str(dataItem.level),
                dataItem.grade,
                useCode,
                useNode

            ]):
                print(value)
                ws1.cell(colIndex + 2, rowIndex + 1).value = value

        else:

            for rowIndex, value in enumerate([
                dataItem.seqNo,
                config.plant_classification,
                config.productionLine_classification,
                dataItem.steelNo,
                dataItem.packageName,
                dataItem.steelType,
                "",
                dataItem.steelLength,
                dataItem.steelWidth,
                dataItem.steelThick,
                dataItem.detectTime,
                dataItem.levelInfoUp,
                dataItem.levelInfoUnder,
                dataItem.levelStr,
                dataItem.levelCode,
                useCode,
                useNode
                # len(dataItem.level) > 0+1,
                # "否" if len(dataItem.level) > 0 else "是",

            ]):
                print(value)
                ws1.cell(colIndex + 2, rowIndex + 1).value = value
                #  流水号	板号	捆包号	钢种	长度	宽度	厚度	检测时间	判级	是否合格	判级原因	人工
    if not excel_path:
        return workbook

    saveDir = Path(excel_path).parent
    if not saveDir.exists():
        saveDir.mkdir(parents=True)

    workbook.save(excel_path)


listDatas = []


def getSaveExcelFileName(data):
    return f"out/{data[0].seqNo}~{data[-1].seqNo}.xlsx"


def append(steelInfo):
    global listDatas
    listDatas.append(steelInfo)
    # if len(listDatas) >= 100:
    #     saveExcel(listDatas, getSaveExcelFileName(listDatas))
    #     listDatas = []

def saveExcel_():
    global listDatas
    saveExcel(listDatas, getSaveExcelFileName(listDatas))
    listDatas = []

# saveExcel([[1,2,3,4],[1,2,3,3]], "test.xlsx")
