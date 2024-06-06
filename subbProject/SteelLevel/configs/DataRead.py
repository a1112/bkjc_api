import requests
from openpyxl.reader.excel import load_workbook

from config import xlsxFile, steelLevelTabelServerUrl


def getSheetData(workbook, name):
    steelSheet = workbook.get_sheet_by_name(name)
    data = []
    for row in steelSheet.iter_rows(values_only=True):
        data.append(row)
    return data


def xlsxDataGet():
    xlsx = xlsxFile
    workbook = load_workbook(xlsx)
    reData = {}
    for name in workbook.get_sheet_names():
        reData[name] = getSheetData(workbook, name)
    return reData


def LevelDataGet():

    if steelLevelTabelServerUrl:
        try:
            return requests.get(steelLevelTabelServerUrl).json()
        except:
            pass
    return xlsxDataGet()