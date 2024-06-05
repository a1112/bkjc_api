import time
from pathlib import Path

from threading import Thread

import openpyxl
from .Base.module.UserDataProject import getUserDataProject
import config
from .DataSet import LevelSet
from openpyxl import load_workbook


class AddUseDataThread(Thread):
    def __init__(self):
        super().__init__()
        inFileFolder = config.get_path("in")
        self.inFileFolder = Path(inFileFolder)
        self.fileCache = []

        self.start()

    def getSheetData(self, steelSheet):
        data = []
        for row in steelSheet.iter_rows(values_only=True):
            data.append(row)
        return data

    def readXlsxData(self,xlsx_file):
        workbook = openpyxl.load_workbook(xlsx_file)
        ws1 = workbook.worksheets[0]
        data = self.getSheetData(ws1)
        titles = data[0]
        datas = data[1:]
        res = []
        for item in datas:
            res_i= {}
            for index, k in enumerate(titles):
                if k is None:
                    k = f"None_{index}"
                res_i[k] = item[index]
            res.append(res_i)
        workbook.close()
        return res

    def run(self):
        while True:
            xlsx_files = self.inFileFolder.glob("*.xlsx")
            for f_ in xlsx_files:
                if f_.name in self.fileCache:
                    continue
                self.fileCache.append(f_.name)
                if LevelSet.hasUserFileData(f_.name):
                    continue
                if "~$" in f_.name:
                    continue
                data = self.readXlsxData(str(f_))
                for item in data:
                    item["fileName"] = f_.name
                    dataProject = getUserDataProject(item)
                    LevelSet.addUserData(dataProject)
            time.sleep(10)


AddUseDataThread()
