from openpyxl.reader.excel import load_workbook

from pathlib import Path

xlsxFolder = Path(fr"C:\project\bkjc_api\out")


def getSheetData(steelSheet):
    data = []
    for row in steelSheet.iter_rows(values_only=True):
        data.append(row)
    return data


rootWork = load_workbook(str("../template/templateDataOut.xlsx"))
root_ws1 = rootWork.get_sheet_by_name("Sheet1")
startRow = 2
okCount = 0
errCount = 0
unKnow = 0


def appendData(dataItem):
    global startRow, root_ws1, okCount, errCount, unKnow
    for rowIndex, item in enumerate(dataItem):
        ls = item[14]
        lc = item[15]
        if lc is None:
            unKnow += 1
        elif ls == lc:
            okCount += 1
        else:
            errCount += 1
        for colIndex, value in enumerate(item):
            root_ws1.cell(rowIndex + startRow, colIndex + 1).value = value
    print(startRow)
    startRow += len(dataItem)


for xmlF in xlsxFolder.glob("*.xlsx"):
    workbook = load_workbook(str(xmlF))
    print(xmlF)
    ws1 = workbook.get_sheet_by_name("Sheet1")
    data = getSheetData(ws1)
    appendData(data[1:])
    # workbook.close()
tIndex = 1
for colIndex, value in enumerate(["一致", "不一致", "无效", "一致率%"]):
    root_ws1.cell(tIndex, colIndex + 20).value = value
tIndex += 1
for colIndex, value in enumerate([okCount, errCount,unKnow, str(int(okCount / (okCount + errCount) * 100))]):
    root_ws1.cell(tIndex, colIndex + 20).value = value
rootWork.save("out2.xlsx")
